import os
import pathlib
from datetime import date
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox

import openpyxl  # Excel
import xlrd  # Excel .xls only
from PIL import Image, ImageTk

os.chdir(os.path.dirname(__file__))


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

class StudentRegistrationSystem(Tk):
    
    def __init__(self):
        super().__init__()
        
        self.title("Student Registration System")
        self.geometry("1200x675+210+50")
        self.resizable(False,False)
        self.config(bg=background)
        
        # Student details
        self.lframe = LabelFrame(
            text="Student's Details",
            font=20,
            bd=2,
            width=900,
            bg=framebg,
            fg=framefg,
            height=250,
            relief=GROOVE)
        self.lframe.place(relx=0.05, rely=0.25)
        
        # Parent details
        self.pframe = LabelFrame(
            text="Parent's Details",
            font=20,
            bd=2,
            width=900,
            bg=framebg,
            fg=framefg,
            height=220,
            relief=GROOVE)
        self.pframe.place(relx=0.05, rely=0.65)
        
        self.frame = Frame(bd=3, bg="black", width=200, height=200, relief=GROOVE)
        self.frame.place(relx=0.82, rely=0.18)
        
        self.img= PhotoImage(file="Images/upload photo.png")
        self.lbl = Label(self.frame, bg="black", image=self.img)
        self.lbl.place(relx=0, rely=0)
        
        #Buttons
        Button(self,text="Upload", width=19, height=2, font="Arial 12 bold", bg="lightblue", command=self.showimage).place(
            relx=0.82,
            rely=0.5)
        self.save_button = Button(
            self,
            text="Save",
            width=19,
            height=2,
            font="Arial 12 bold", 
            bg="lightgreen",
            command=self.Save)
        self.save_button.place(relx=0.82, rely=0.6)
        Button(self,text="Reset", width=19, height=2, font="Arial 12 bold", bg="lightpink", command=self.Clear).place(
            relx=0.82,
            rely=0.7)
        Button(self,text="Exit", width=19, height=2, font="Arial 12 bold", bg="grey", command=self.exit).place(
            relx=0.82,
            rely=0.8)
        
        self.__create_student_base()
        self.__create_top_widgets()
        self.__create_date_registration()
        self.__create_student_details()
        self.__create_parents_details()
        
    def exit(self, event=None):
        answer = messagebox.askokcancel('Выход', 'Вы точно хотите выйти?')
        if answer:
            self.quit()
    
    def registration_no(self):
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        row = sheet.max_row
        max_row_value = sheet.cell(row=row, column=1).value
        print(max_row_value, type(max_row_value))
        try:
            self.Registration.set(int(max_row_value)+1)
        except Exception:
            self.Registration.set("1")

    def Clear(self):
        global img
        self.Name.set('')
        self.DOB.set('')
        self.Religion.set('')
        self.Skills.set('')
        self.Ocupation.set('')
        self.father_name.set('')
        self.mother_name.set('')
        self.M_Ocupation.set('')
        self.Class.set('Select Class')
        self.radio.set(0)
        
        self.registration_no()
        self.save_button.config(state='normal')
        
        img1 = PhotoImage(file="Images\\upload photo.png")
        self.lbl.config(image=img1)
        self.lbl.image = img1
        
        img=""
    
    def Save(self):
        self.R1 = self.Registration.get()
        N1 = self.Name.get()
        C1 = self.Class.get()
        try:
            G1 = self.gender
            if self.radio.get() == 0:
                messagebox.showerror("Error", "Select Gender!")
                return
        except:
            messagebox.showerror("Error", "Select Gender!")
            return
        D2 = self.DOB.get()
        D1 = self.Date.get()
        Re1 = self.Religion.get()
        S1 = self.Skills.get()
        fathername = self.father_name.get()
        mothername = self.mother_name.get()
        F1 = self.Ocupation.get()
        M1 = self.M_Ocupation.get()
        
        if N1 == '' or C1 == "Select Class" or D2 == "" or S1 =="" or Re1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
            messagebox.showerror("Error", "Few Data is missing")
        else:
            file =  openpyxl.load_workbook("Student_data.xlsx")
            sheet = file.active
            sheet.cell(column=1,row=sheet.max_row+1, value=self.R1)
            sheet.cell(column=2,row=sheet.max_row, value=N1)
            sheet.cell(column=3,row=sheet.max_row, value=C1)
            sheet.cell(column=4,row=sheet.max_row, value=G1)
            sheet.cell(column=5,row=sheet.max_row, value=D2)
            sheet.cell(column=6,row=sheet.max_row, value=D1)
            sheet.cell(column=7,row=sheet.max_row, value=Re1)
            sheet.cell(column=8,row=sheet.max_row, value=S1)
            sheet.cell(column=9,row=sheet.max_row, value=fathername)
            sheet.cell(column=10,row=sheet.max_row, value=mothername)
            sheet.cell(column=11,row=sheet.max_row, value=F1)
            sheet.cell(column=12,row=sheet.max_row, value=M1)
            
            file.save(r'Student_data.xlsx')
            try:
                img.save(imagetype) #img.save("Student_Images/"+str(self.R1)+".jpg")
            except:
                messagebox.showinfo("Info", "Profile Photo is not available!")
            
            messagebox.showinfo("Info", "Sucessfully data entered!")
            
            self.Clear()
            self.registration_no()
            
                    
    def showimage(self):
        global fileimage
        global img
        global imagetype
        fileimage = filedialog.askopenfilename(
            initialdir=os.getcwd(), 
            title="Select image file", 
            filetypes=(("JPG File", "*.jpg"),("PNG File","*.png"),("All files","*.*"))
            )
        img = (Image.open(fileimage))
        resize_img = img.resize((190,190))
        photo2 = ImageTk.PhotoImage(resize_img)
        self.lbl.config(image=photo2)
        self.lbl.image = photo2
        imagetype = fileimage
        
    def __create_student_details(self):
        full_name = Label(self.lframe, text="Full name:", font="Arial 13", bg=framebg, fg=framefg)
        full_name.place(relx=0.06, rely=0.11)
        self.Name = StringVar()
        name_entry = Entry(self.lframe, textvariable=self.Name, width=20, font="Arial 10")
        name_entry.place(relx=0.19, rely=0.11)
        
        self.DOB = StringVar()
        DOB_entry = Entry(self.lframe, textvariable=self.DOB, width=20, font="Arial 10")
        DOB_entry.place(relx=0.19, rely=0.3)
        date_birth = Label(self.lframe, text="Date of Birth:", font="Arial 13", bg=framebg, fg=framefg)
        date_birth.place(relx=0.06, rely=0.3)
        
        self.radio = IntVar()
        self.R1 = Radiobutton(
            self.lframe, 
            text="Male",
            variable=self.radio, 
            value=1, 
            bg=framebg, 
            fg=framefg,
            command=self.selection)
        self.R1.place(relx=0.18, rely=0.5)
        
        self.R2 = Radiobutton(
            self.lframe,
            text="Female", 
            variable=self.radio,
            value=2, 
            bg=framebg, 
            fg=framefg, 
            command=self.selection)
        
        self.R2.place(relx=0.27, rely=0.5)
        
        wgender = Label(self.lframe, text="Gender:", font="Arial 13", bg=framebg, fg=framefg)
        wgender.place(relx=0.06, rely=0.5)
        
        self.Class = Combobox(
            self.lframe, 
            values=['1','2','3','4','5','6','7','8','9','10','11'], 
            font="Roboto 10",
            width=17,
            state="r")
        self.Class.place(relx=0.65, rely=0.11)
        self.Class.set("Select Class")
        WClass = Label(self.lframe, text="Class:", font="Arial 13", bg=framebg, fg=framefg)
        WClass.place(relx=0.45, rely=0.11)
        
        self.Religion = StringVar()
        religion_entry = Entry(self.lframe, textvariable=self.Religion, width=20, font="Arial 10")
        religion_entry.place(relx=0.65, rely=0.31)
        WReligion = Label(self.lframe, text="Religion:", font="Arial 13", bg=framebg, fg=framefg)
        WReligion.place(relx=0.45, rely=0.313)
        
        self.Skills = StringVar()
        skills_entry = Entry(self.lframe, textvariable=self.Skills, width=20, font="Arial 10")
        skills_entry.place(relx=0.65, rely=0.5)
        WSkills = Label(self.lframe, text="Skills:", font="Arial 13", bg=framebg, fg=framefg)
        WSkills.place(relx=0.45, rely=0.5)
    
    def __create_parents_details(self):
        self.father_name = StringVar()
        self.fname_entry = Entry(self.pframe, textvariable=self.father_name, width=20, font="Arial 10")
        self.fname_entry.place(relx=0.2, rely=0.15)
        self.Wfather_name = Label(
            self.pframe,
            text="Father's Name:",
            font="Arial 13", 
            bg=framebg,
            fg=framefg)
        self.Wfather_name.place(relx=0.05, rely=0.15)
        
        self.Ocupation = StringVar()
        self.ocupation_entry = Entry(self.pframe, textvariable=self.Ocupation, width=20, font="Arial 10")
        self.ocupation_entry.place(relx=0.2, rely=0.42)
        self.WOcupation = Label(self.pframe, text="Occupation:", font="Arial 13", bg=framebg, fg=framefg)
        self.WOcupation.place(relx=0.05, rely=0.42)
        
        self.mother_name = StringVar()
        self.mname_entry = Entry(self.pframe, textvariable=self.mother_name, width=20, font="Arial 10")
        self.mname_entry.place(relx=0.65, rely=0.15)
        self.Wmother_name = Label(
            self.pframe,
            text="Mother's Name:", 
            font="Arial 13", 
            bg=framebg, 
            fg=framefg)
        self.Wmother_name.place(relx=0.5, rely=0.15)
        
        self.M_Ocupation = StringVar()
        self.m_ocupation_entry = Entry(self.pframe, textvariable=self.M_Ocupation, width=20, font="Arial 10")
        self.m_ocupation_entry.place(relx=0.65, rely=0.42)
        self.Wmocupation = Label(self.pframe, text="Occupation:", font="Arial 13", bg=framebg, fg=framefg)
        self.Wmocupation.place(relx=0.5, rely=0.42)
        
    # Top widgets
    def __create_top_widgets(self):
        self.label = Label(
            text="STUDENT REGISTRATION\t\t",
            width=12,
            height=3,
            bg="#c36464",
            fg="#fff",
            font="Arial 18 bold",
            )
        self.label.pack(fill=X)
        
        self.Search = StringVar()
        self.enry = Entry(textvariable=self.Search, 
                width=12,
                bd=2,
                font="arial 20")
        self.enry.place(relx=0.64,rely=0.05)
    
        # search box to update
        self.searchimage = PhotoImage(file="Images/search.png")
        self.search_button = Button(
            text="Search",
            compound=LEFT,
            width=123,
            height=2,
            bd=2,
            bg="#68ddfa",
            fg="black",
            image=self.searchimage,
            font="Arial 13 bold")
        self.search_button.place(relx=0.82, rely=0.04, relheight=0.07)       
    
        self.layerimg = PhotoImage(file="Images/Layer 4.png")
        self.update_button = Button( 
            bg="#c36464",
            image=self.layerimg) 
        self.update_button.place(relx=0.1, rely=0.04)    
    
    # Date and Registration
    def __create_date_registration(self):
        self.lreg = Label(
            text="Registration No:", 
            font="Arial 13",
            bg = "#06283D", 
            fg="#EDEDED")
        self.lreg.place(relx=0.04, rely=0.18)
        self.ldate = Label(
            text="Date:", 
            font="Arial 13",
            bg = "#06283D", 
            fg="#EDEDED")
        self.ldate.place(relx=0.4, rely=0.18)
        
        self.Registration = StringVar()
        self.Date = StringVar()
        
        self.reg_entry = Entry(
            textvariable=self.Registration,
            width=14,
            font="Arial 10"
        )
        self.reg_entry.place(relx=0.15, rely=0.18)
        
        self.registration_no()
        
        # Get todays date
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        self.data_entry = Entry(
            textvariable=self.Date,
            width=15,
            font="Arial 10"
        )
        self.data_entry.place(relx=0.44, rely=0.18)
        
        self.Date.set(d1)
        
    def __create_student_base(self):
        self.student_data = pathlib.Path("Student_data.xlsx")
        if self.student_data.exists():
            pass
        else: 
            file = openpyxl.Workbook()
            sheet = file.active
            sheet['A1'] = "Registration No."
            sheet['B1'] = "Name"
            sheet['C1'] = "Class"
            sheet['D1'] = "Gender"
            sheet['E1'] = "DOB"
            sheet['F1'] = "Date of Registration"
            sheet['G1'] = "Regilion"
            sheet['H1'] = "SKill"
            sheet['I1'] = "Father Name"
            sheet['J1'] = "Mother Name"
            sheet['K1'] = "Father's Occupation"
            sheet['L1'] = "Mother's Occupation"             

            self.student_data.save('Student_data.xlsx')
    
    def selection(self):
        value = self.radio.get()
        if value==1:
            self.gender = "Male"
        else:
            self.gender = "Female"     




if __name__ == "__main__":
    MyApp = StudentRegistrationSystem()
    MyApp.mainloop()